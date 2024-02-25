from openpyxl import load_workbook
import csv
import kacplansheet as ks

class KacWorkBook:
    def __init__(self, conf, filename):
        self.conf = conf
        self.filename = filename
        self.workbook = load_workbook(filename=self.filename, data_only=True)
        #sheet = workbook.active
        self.sheet_names = self.workbook.sheetnames
        self.summary = self.sheet_names.pop(0)
    
    def show_sheet_list(self):
        if self.conf["verbose"]:
            print("[show_sheet_list]")
        sheet_names =  self.sheet_names

        for sheet_name in sheet_names:
            print(sheet_name)

    def get_sheet_list(self, id):
        sheet = ks.KacPlanSheet(self.conf, self.workbook[id], id)
        sheet.get_month_block()
        #sheet.print_plan_list()
        return sheet.get_plan_list()

    def output_csv(self, filename):
        sheet_names =  self.sheet_names

        plan_list = []
        for sheet_name in sheet_names:
            plan_list.extend(self.get_sheet_list(sheet_name))
        
        header = ['date', 'weekday', 'time_symbol', 'place', 'id']
        # CSVファイルにデータを書き込む
        with open(filename, 'w', newline='') as file:
            #writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC)
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)
            writer.writerow(header)
            writer.writerows(plan_list)

    def get_sheet_lists(self):
        sheet_names =  self.sheet_names

        plan_list = []
        for sheet_name in sheet_names:
            plan_list.extend(self.get_sheet_list(sheet_name))
        
        print("result:")
        for plan in plan_list:
            print(plan)
            